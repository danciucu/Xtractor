import os
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

import globalvars


class field_notes():
    def __init__(self) -> None:
        pass


    def create(self, dictionary):
        # array for keywords that the loop will skip
        skip_keywords = ['Total Quantity', 'CS1', 'CS2', 'CS3', 'CS4', 'Description', 'Work Items Action', 'Work Items Priority', 'Work Items Responsibility']
        # Define the color fills
        fill_green = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        fill_orange = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        # Create the rules
        rule_green = FormulaRule(formula=['$D19="CS1"'], fill=fill_green)
        rule_yellow = FormulaRule(formula=['$D19="CS2"'], fill=fill_yellow)
        rule_orange = FormulaRule(formula=['$D19="CS3"'], fill=fill_orange)
        rule_red = FormulaRule(formula=['$D19="CS4"'], fill=fill_red)

        try:
            excel_path = 'C:/Users/' + globalvars.user_path + '/AECOM/KYTC NBIS Inspections - 2023-2025/400_Technical/200_Templates/LA ## D## - County/1_AECOM Bridges/Incomplete/Field Note Spreadsheets/MACRO_Inspection Element Library_SNBI.xlsx'
            excel_template = r'%s' % excel_path
            # create a new excel file
            wb = openpyxl.load_workbook(excel_template)
            #
            referece_sheet = wb['Reference - DO NOT DELETE']
            # main tab
            sheet = wb['Info, NBI, Work']
            # sheet to copy
            sheet_to_copy = wb['# - Element Temp']
            # populate the cells
            
            for keyword in dictionary:
                if keyword == 'Work Items Description':
                    for k in range(len(dictionary[keyword][1])):
                        # add working items into the first tab
                        cell_obj0_wi = sheet.cell(row = dictionary['Work Items Description'][0][k][0], column = dictionary['Work Items Description'][0][k][1], value = dictionary['Work Items Description'][1][k])
                        cell_obj1_wi = sheet.cell(row = dictionary['Work Items Action'][0][k][0], column = dictionary['Work Items Action'][0][k][1], value = dictionary['Work Items Action'][1][k])
                        cell_obj2_wi = sheet.cell(row = dictionary['Work Items Priority'][0][k][0], column = dictionary['Work Items Priority'][0][k][1], value = dictionary['Work Items Priority'][1][k])
                        cell_obj3_wi = sheet.cell(row = dictionary['Work Items Responsibility'][0][k][0], column = dictionary['Work Items Responsibility'][0][k][1], value = dictionary['Work Items Responsibility'][1][k])

                elif keyword == 'Elements':
                    for j in range(len(dictionary[keyword][1])):
                        # add the notes into the first tab
                        cell_obj0 = sheet.cell(row = dictionary['Elements'][0][j][0], column = dictionary['Elements'][0][j][1], value = int(dictionary['Elements'][1][j]))
                        cell_obj1 = sheet.cell(row = dictionary['Total Quantity'][0][j][0], column = dictionary['Total Quantity'][0][j][1], value = float(dictionary['Total Quantity'][1][j]))
                        cell_obj2 = sheet.cell(row = dictionary['CS1'][0][j][0], column = dictionary['CS1'][0][j][1], value = dictionary['CS1'][1][j])
                        cell_obj3 = sheet.cell(row = dictionary['CS2'][0][j][0], column = dictionary['CS2'][0][j][1], value = dictionary['CS2'][1][j])
                        cell_obj4 = sheet.cell(row = dictionary['CS3'][0][j][0], column = dictionary['CS3'][0][j][1], value = dictionary['CS3'][1][j])
                        cell_obj5 = sheet.cell(row = dictionary['CS4'][0][j][0], column = dictionary['CS4'][0][j][1], value = dictionary['CS4'][1][j])
                        cell_obj6 = sheet.cell(row = dictionary['Description'][0][j][0], column = dictionary['Description'][0][j][1], value = dictionary['Description'][1][j])  
                        # create new tabs with the elements number
                        new_sheet = wb.copy_worksheet(sheet_to_copy)
                        new_sheet.title = dictionary['Elements'][1][j]
                        wb.move_sheet(new_sheet, offset = -2)
                        # add to the tab the informations
                        cell_obh0 = new_sheet.cell(row = 2, column = 1, value = int(dictionary['Elements'][1][j]))
                        cell_obh1 = new_sheet.cell(row = 4, column = 4, value = float(dictionary['Total Quantity'][1][j]))
                        cell_obh2 = new_sheet.cell(row = 5, column = 4, value = float(dictionary['Total Quantity'][1][j]))
                        cell_obh3 = new_sheet.cell(row = 4, column = 6, value = dictionary['CS1'][1][j])
                        cell_obh4 = new_sheet.cell(row = 4, column = 7, value = dictionary['CS2'][1][j])
                        cell_obh5 = new_sheet.cell(row = 4, column = 8, value = dictionary['CS3'][1][j])
                        cell_obh6 = new_sheet.cell(row = 4, column = 9, value = dictionary['CS4'][1][j])
                        cell_obh7 = new_sheet.cell(row = 17, column = 1, value = dictionary['Description'][1][j])
                        # create condition data validation
                        self.create_dataValidation(excel_sheet=new_sheet, input_list="CS1, CS2, CS3, CS4", location='D19:D61')
                        # Apply the rules to the entire range 'D19:D61'
                        new_sheet.conditional_formatting.add('D19:D61', rule_green)
                        new_sheet.conditional_formatting.add('D19:D61', rule_yellow)
                        new_sheet.conditional_formatting.add('D19:D61', rule_orange)
                        new_sheet.conditional_formatting.add('D19:D61', rule_red)

                elif keyword == 'Posting Values':
                    try:
                        for l in range(len(dictionary[keyword][0])):
                            # add posting into the first tab
                            cell_obj0_p = sheet.cell(row = dictionary['Posting Values'][0][l][0], column = dictionary['Posting Values'][0][l][1], value = dictionary['Posting Values'][1][l])
                    except Exception as e:
                        print("An error occurred:", e)

                elif keyword in skip_keywords:
                    continue

                else:
                    cell_obj_main = sheet.cell(row = dictionary[keyword][0][0], column = dictionary[keyword][0][1], value = dictionary[keyword][1])
            #
            
            # save the file
            path = os.path.join(globalvars.save_path, dictionary['Structure ID'][1] + ' Field Notes '+ globalvars.inspection_date[0:4] + globalvars.inspection_date[5:7] +'DD.xlsx')
            wb.save(path)

            #print(wb.sheetnames)
            # add QTL Data Validation
            qtl_dataValidation = self.create_dataValidation(excel_sheet=sheet, reference_sheet=referece_sheet, cell_range_to_copy='$AG$18:$AG$23', location='B2')
            # add Assistant Data Validation
            assistant_dataValidation = self.create_dataValidation(excel_sheet=sheet, reference_sheet=referece_sheet, cell_range_to_copy='$AH$18:$AH$31', location='B3')
            # add Inspection Type Data Validation
            inspection_type_dataValidation = self.create_dataValidation(excel_sheet=sheet, input_list="select one->, Initial, Routine, NSTM, Damage, In-Depth, Special, Service, Scour Monitoring", location='B5')
            # add Route Log Direction Data Validation
            route_log_direction_dataValidation = self.create_dataValidation(excel_sheet=sheet, input_list="select one->, north, east, south, west", location='B8')
            # add Is Insp Consistent with Route Log Data Validation
            is_insp_consistent_route_log_dataValidation = self.create_dataValidation(excel_sheet=sheet, input_list="select one->, Yes, No", location='B9')
            # add Required Posting Data Validation
            posting_dataValidation = self.create_dataValidation(excel_sheet=sheet, input_list="select one->, Yes, No", location='E9')
            # add Posted Appropriately Data Validation
            posted_appropriately_dataValidation = self.create_dataValidation(excel_sheet=sheet, input_list="select one->, Yes, No", location='E9')
            # add Non-Cardinal Data Validaton
            non_cardinal_dataValidation = self.create_dataValidation(excel_sheet=sheet, input_list="select one->, Yes, No", location='F11')
            # add Upstream Data Validation
            upstream_dataValidation = self.create_dataValidation(excel_sheet=sheet, input_list="select one->, left, right, N/A", location='E11')
            # add Bridge/Culvert Data Validation
            bridge_culvert_dataValidation = self.create_dataValidation(excel_sheet=sheet, input_list="select one->, Bridge, Culvert", location='B14')
            # add Equipment Data Validation
            equipment_used1_dataValidation = self.create_dataValidation(excel_sheet=sheet, reference_sheet=referece_sheet, cell_range_to_copy='$Q$16:$Q$22', location='B19')
            equipment_used2_dataValidation = self.create_dataValidation(excel_sheet=sheet, reference_sheet=referece_sheet, cell_range_to_copy='$Q$16:$Q$22', location='B20')
            equipment_used3_dataValidation = self.create_dataValidation(excel_sheet=sheet, reference_sheet=referece_sheet, cell_range_to_copy='$Q$16:$Q$22', location='B21')
            # add Scour Observed Data Validation
            scour_prev_dataValidation = self.create_dataValidation(excel_sheet=sheet, input_list="Prev. Rating->, No Scour, Minor Scour, Moderate Scour, Major Scour, N/A", location='E20')
            scour_current_dataValidation = self.create_dataValidation(excel_sheet=sheet, input_list="Prev. Rating->, No Scour, Minor Scour, Moderate Scour, Major Scour, N/A", location='F20')
            scour_rating_dataValidation = self.create_dataValidation(excel_sheet=sheet, reference_sheet=referece_sheet, cell_range_to_copy='$AV$68:$AV$79', location='E21')
            # add Deck Ratings Data Validation
            deck_prev_dataValidation = self.create_dataValidation(excel_sheet=sheet, reference_sheet=referece_sheet, cell_range_to_copy='$AN$3:$AN$14', location='B25')
            deck_current_dataValidation = self.create_dataValidation(excel_sheet=sheet, reference_sheet=referece_sheet, cell_range_to_copy='$AN$3:$AN$14', location='C25')
            # add Superstructure Data Validation
            superstructure_prev_dataValidation = self.create_dataValidation(excel_sheet=sheet, reference_sheet=referece_sheet, cell_range_to_copy='$AN$3:$AN$14', location='B28')
            superstructure_current_dataValidation = self.create_dataValidation(excel_sheet=sheet, reference_sheet=referece_sheet, cell_range_to_copy='$AN$3:$AN$14', location='C28')
            # add Substructure Data Validation
            substructure_prev_dataValidation = self.create_dataValidation(excel_sheet=sheet, reference_sheet=referece_sheet, cell_range_to_copy='$AN$3:$AN$14', location='B31')
            substructure_current_dataValidation = self.create_dataValidation(excel_sheet=sheet, reference_sheet=referece_sheet, cell_range_to_copy='$AN$3:$AN$14', location='C31')
            # add Culvert Data Validation
            culvert_prev_dataValidation = self.create_dataValidation(excel_sheet=sheet, reference_sheet=referece_sheet, cell_range_to_copy='$AN$3:$AN$14', location='B34')
            culvert_current_dataValidation = self.create_dataValidation(excel_sheet=sheet, reference_sheet=referece_sheet, cell_range_to_copy='$AN$3:$AN$14', location='C34')
            # add Rail Data Validation
            rail_current_dataValidation = self.create_dataValidation(excel_sheet=sheet, reference_sheet=referece_sheet, cell_range_to_copy='$AN$3:$AN$14', location='C37')
            # add Rail Transition Data Validation
            rail_transition_current_dataValidation = self.create_dataValidation(excel_sheet=sheet, reference_sheet=referece_sheet, cell_range_to_copy='$AN$3:$AN$14', location='C40')
            # add Bearing Data Validation
            bearing_current_dataValidation = self.create_dataValidation(excel_sheet=sheet, reference_sheet=referece_sheet, cell_range_to_copy='$AN$3:$AN$14', location='C43')
            # add Joints Data Validation
            joint_current_dataValidation = self.create_dataValidation(excel_sheet=sheet, reference_sheet=referece_sheet, cell_range_to_copy='$AN$3:$AN$14', location='C46')
            # add Channel Data Validation
            channel_prev_dataValidation = self.create_dataValidation(excel_sheet=sheet, reference_sheet=referece_sheet, cell_range_to_copy='$AN$3:$AN$14', location='B49')
            channel_current_dataValidation = self.create_dataValidation(excel_sheet=sheet, reference_sheet=referece_sheet, cell_range_to_copy='$AN$3:$AN$14', location='C49')
            # add Channel Data Validation
            channel_protection_current_dataValidation = self.create_dataValidation(excel_sheet=sheet, reference_sheet=referece_sheet, cell_range_to_copy='$AN$3:$AN$14', location='C52')  
            # add Work Items Data Validation
            work_items_action_dataValidation = self.create_dataValidation(excel_sheet=sheet, reference_sheet=referece_sheet, cell_range_to_copy='$AJ$19:$AJ$65', location='E57:E69')
            work_items_comment_dataValidation = self.create_dataValidation(excel_sheet=sheet, reference_sheet=referece_sheet, cell_range_to_copy='$AK$17:$AK$20', location='I57:I69')
            # get the saved file
            #excel = ExcelCompiler(filename=path)
            for i in range(len(dictionary['Elements'][1])):
                # get the element tab
                sheet = dictionary['Elements'][1][i]
                # create defects data validation
                
                # set it as active sheet
                ws = wb[sheet]
                self.create_dataValidation(excel_sheet=ws, cell_range_to_copy='$C$6:$C$15', location='C19:C61')
                # get all the defects in C6:C15 cells
                #defects = ','.join([str(excel.evaluate(f'{sheet}!C{i}')) for i in range(6, 15)])
                # create defects data validation
                #dv2 = DataValidation(type='list', formula1=f'"{defects}"')
                #ws.add_data_validation(dv2)
                #dv2.add('C19:C61')
            # save the file again
            wb.save(path)

        except Exception as e:
            print("An error occurred:", e)
        finally:
            # close the workbook
            if 'wb' in locals():
                wb.close()



    def create_dataValidation(self, excel_sheet, location, input_list=None, cell_range_to_copy=None, reference_sheet=None):
        # If the user inputs a list, set the formula accordingly
        if input_list:
            formula = f'"{input_list}"'
        # If the user inputs a range to copy, set the formula accordingly
        elif cell_range_to_copy:
            # If the user points to a reference_sheet, set the sheet to copy accordingly
            if reference_sheet:
                sheet_used = reference_sheet
            else:
                sheet_used = excel_sheet

            # Ensure proper format for the formula
            formula = f"{quote_sheetname(sheet_used.title)}!{cell_range_to_copy}"

        # Print the formula to debug
        #print(f"Data Validation Formula: {formula}")

        # Create and apply the data validation
        dv = DataValidation(type='list', formula1=formula)
        excel_sheet.add_data_validation(dv)
        dv.add(location)