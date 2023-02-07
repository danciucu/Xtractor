import os
import openpyxl

class Xtractor():
    def __init__(self):
        super().__init__()
        # set the path for the txt files
        self.dirpath = 'path'
        # set the path for the excel files that will be generated
        self.savepath = 'path'
        # create a new excel spreadsheet
        self.generate_excel(self.dirpath, self.savepath)

    def generate_excel(self, dirpath, savepath):
        # create empty string variables
        filename = ''
        filepath = ''
        notes = ''
        elements = ''
        qunatities = ''
        excel_template = 'file'
        # get the number of .txt files are in the folder
        file_numbers = len(os.listdir(dirpath))
        # loop over the entire number of txt files available
        for i in range(file_numbers):
            # create filepath
            filepath = self.importfile(dirpath, i)
            # get the filename
            filename = self.get_filename(dirpath, i)
            # get inspection notes
            notes = self.inspection_notes(filepath)
            # get elements' number
            elements = self.element_number(notes)
            # get quantities
            qunatities = self.element_quantities(notes)
            # create a new excel spreadsheet
            self.new_excel(excel_template, savepath, filename, elements, qunatities)
        

    # function that returns the path of the txt file
    def importfile(self, directory, index):
        # get the folder's path
        file = os.listdir(directory)
        # join the folder's path with the name of the file
        path = os.path.join(directory, file[index])
        return path            

    # function that returns the bridge ID
    def get_filename(self, directory, index):
        # get the folder's path
        file = os.listdir(directory)
        # get the file name
        filename = file[index][len(file[index]) - 25:len(file[index]) - 15]  
        return filename         

    def inspection_notes(self, filepath):
        # create an empty array
        elements = []
        # open the txt file
        file = open(filepath, 'r')
        # read the lines of the file
        lines = file.readlines()
        # create a variable that loops over the lines
        lineofinterest = 0
        # create a variable to monitor the empty lines
        breaklines = 0
        # create a variable to loop over the lines of interest
        i = 0

        # define a loop that goes over all the lines in the txt file
        while True:
            # update variable
            lineofinterest += 1
            # if the line contains what the user is looking for
            if 'ELEM   ELEMENT NAME' in lines[lineofinterest]:
                # loop over the lines below the line that contains the wanted string
                while i < len(lines) - lineofinterest - 1:
                    # update variable
                    i += 1
                    # if the line is empty
                    if len(lines[lineofinterest + i]) < 2:
                        # update variable
                        breaklines += 1
                        # if there were two empty lines from the beginning
                        if breaklines == 2:
                            # update variables and stop
                            breaklines = 0
                            i = 0
                            break
                    # if the line contains 'Inspection' jump 8 lines
                    if 'Inspection' in lines[lineofinterest + i]:
                        i += 7
                        continue
                    # if the line is not empty add it to the elements array
                    if len(lines[lineofinterest + i]) > 2:
                        elements.append(lines[lineofinterest + i])
            # if the line contains 'Work Candidates Report' then stop the process
            if 'Work Candidates Report' in lines[lineofinterest]:
                # close txt file
                file.close()
                return elements

    # function that extracts the inspection notes
    def element_number(self, notes):
        # create empty variables
        digits = 0
        first_digit = 0
        last_digit = 0
        # create empty array
        elements = []
    
        # loop over all the sentences/strings in the notes
        for strings in notes:
            # loop over the charactes of each string
            for i in range(len(strings)):
                # if the character is a digit and the number of digits is null
                if strings[i].isdigit() == True and digits == 0:
                    # update first_digit and the digits count
                    first_digit = i
                    digits += 1
                # if the character is a digit and the next character is not
                if strings[i].isdigit() == True and strings[i + 1].isdigit() == False:
                    # update last_digit and then stop
                    last_digit = i
                    break
            # if the lenght of the string from first_digit to last_digit + 1 is greater than 1 and the first_digit is within the first 10 chars
            if len(strings[first_digit:last_digit + 1]) > 1 and len(strings) > 35 and first_digit < 6 :
                # update the elements array
                elements.append(strings[first_digit:last_digit + 1])

            # reset varables
            digits = 0
            first_digit = 0
            last_digit = 0

        return elements

    # function that extracts the element quantities and the description
    def element_quantities(self, notes):
        # create empty variables
        slashes = 0
        last_slash = 0
        count = 0
        # create empty arrays
        strings_main = []
        quantities = []
        total_qunatities = []
        cs1_quantities = []
        cs2_quantities = []
        cs3_quantities = []
        cs4_quantities = []
        description = []
        strings_description = []
        # create an empty string
        temp = ''

        # loop over the strings in the notes
        for strings in notes:
            # update count
            count += 1
            # if there is a slash followed by two digits inside of the string
            if '.00' in strings[len(strings) - 6 : len(strings) - 1]:
                # remember the key of that string
                strings_main.append(count)
                # loop over the characters in that string
                for i in range(len(strings)):
                    # if there is a slash followed by a number
                    if strings[i] == "/" and strings[i + 1].isdigit() == True:
                        # update slashes variable
                        slashes += 1
                        # if there are two slahses found
                        if slashes == 2:
                        # remember the location of last slash and stop
                            last_slash = i
                            break
                # add the string.split() from last_slash to quantities array 
                quantities.append(strings[last_slash:].split())
            # if there ar no slashes
            else:
                # remember the key of that string
                strings_description.append(count)
        
        # reset variables
        slashes = 0
        last_slash = 0

        count = 0

        # loop over splitted values of the quantities
        for quantity in quantities:
            # update total quantities
            total_qunatities.append(quantity[1])
            # check the lenght of the array of strings
            try:
                cs1_quantities.append(quantity[len(quantity) - 4])
            except:
                cs1_quantities.append('?')
            try:
                cs2_quantities.append(quantity[len(quantity) - 3])
            except:
                cs2_quantities.append('?')
            try:
                cs3_quantities.append(quantity[len(quantity) - 2])
            except:
                cs4_quantities.append('?')
            try:
                cs4_quantities.append(quantity[len(quantity) - 1])
            except:
                cs4_quantities.append('?')

        # loop over all stings are not comments
        for i in range(len(strings_main)):
            if i == len(strings_main) - 1:
                difference = 1
            else:
                difference = int(strings_main[i + 1] - strings_main[i])
            for j in range(1, difference):
                temp += notes[int(strings_main[i] + j - 1)]
                #print('temp = ', temp)
            description.append(temp.replace("\n", ''))
            temp = ''

        return [total_qunatities, cs1_quantities, cs2_quantities, cs3_quantities, cs4_quantities, description]

    def new_excel(self, excel_template, savepath, filename, elements, quantities):
        # total number of elements per txt file
        no_elements = len(elements)
        # create a new excel file
        wb = openpyxl.load_workbook(excel_template)
        sheet = wb['Info, NBI, Work']
        sheet_to_copy = wb['# - Element Temp']
        # loop over all the elements extracted
        for i in range(no_elements):
            # add the notes into the first tab
            cell_obj0 = sheet.cell(row = 84 + i, column = 1, value = elements[i])
            cell_obj1 = sheet.cell(row = 84 + i, column = 2, value = quantities[0][i])
            cell_obj2 = sheet.cell(row = 84 + i, column = 3, value = quantities[1][i])
            cell_obj3 = sheet.cell(row = 84 + i, column = 4, value = quantities[2][i])
            cell_obj4 = sheet.cell(row = 84 + i, column = 5, value = quantities[3][i])
            cell_obj5 = sheet.cell(row = 84 + i, column = 6, value = quantities[4][i])
            cell_obj6 = sheet.cell(row = 84 + i, column = 7, value = quantities[5][i])                
            # create new tabs with the elements number
            new_sheet = wb.copy_worksheet(sheet_to_copy)
            new_sheet.title = elements[i]
            wb.move_sheet(new_sheet, offset = -2)
            # add to the tab the informations
            cell_obh0 = new_sheet.cell(row = 2, column = 1, value = int(elements[i]))
            cell_obh1 = new_sheet.cell(row = 4, column = 4, value = quantities[0][i])
            cell_obh2 = new_sheet.cell(row = 4, column = 6, value = quantities[1][i])
            cell_obh3 = new_sheet.cell(row = 4, column = 7, value = quantities[2][i])
            cell_obh4 = new_sheet.cell(row = 4, column = 8, value = quantities[3][i])
            cell_obh5 = new_sheet.cell(row = 4, column = 9, value = quantities[4][i])
            cell_obh6 = new_sheet.cell(row = 17, column = 1, value = quantities[5][i])
        # save the file
        wb.save(os.path.join(savepath, filename + ' Field Notes 2023MM.xlsx'))
        wb.close()            
            
if __name__ == "__main__":
    app = Xtractor()
