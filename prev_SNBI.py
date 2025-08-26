import os
import openpyxl as x1
from openpyxl import load_workbook
from openpyxl.styles.fills import PatternFill
from openpyxl.styles import Font, colors
import shutil
import math
import globalvars#, folders
from selenium.webdriver.common.by import By
import time

def get_prev_SNBI(driver, bridge_dict, i):
    globalvars.init()

    # open previous SNBI sheet from "Previous Reports" folder
    parent_path = os.path.dirname(globalvars.save_path)
    prev_reports_folder = str(parent_path) + '/' + str(globalvars.bridgeID[i]) + '/Previous Reports/'
    
    xlsx_files = [f for f in os.listdir(prev_reports_folder) if f.endswith('.xlsx')]
    
    if not xlsx_files:
        print('No .xlsx files found in the ' + str(globalvars.bridgeID[i]) + ' Previous Reports folder.')
        return None
    
    if len(xlsx_files) > 1:
        raise Exception(f"Expected exactly one .xlsx file, found {len(xlsx_files)}.")
    
    xlsx_path = os.path.join(prev_reports_folder, xlsx_files[0])
    workbook = load_workbook(xlsx_path, data_only=True) # data_only=True to get cell values, not formulas
    sheet = workbook.active  # get the first (active) sheet

    #bridge_dict['B.SP.03 (Number of Beams)'][1] = sheet["C29"].value
    bridge_dict['B.G.01 (NBIS Length)'][1] = sheet["C51"].value
    bridge_dict['B.G.04 (Minimum Span Length)'][1] = sheet["C54"].value
    bridge_dict['B.G.13 (Maximum Bridge Height)'][1] = sheet["C63"].value
    bridge_dict['B.C.05 (Bridge Railing)'][1] = sheet["C147"].value
    bridge_dict['B.C.06 (Bridge Railing Transition)'][1] = sheet["C148"].value
    bridge_dict['B.C.07 (Bridge Bearing)'][1] = sheet["C149"].value
    bridge_dict['B.C.08 (Bridge Joints)'][1] = sheet["C150"].value
    bridge_dict['B.C.10 (Channel Protection)'][1] = sheet["C152"].value
    bridge_dict['B.C.11 (Scour)'][1] = sheet["C153"].value
